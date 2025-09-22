import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def json_to_excel(json_file_path, excel_file_path):
    """
    将JSON文件中的数据转换为Excel表格
    
    Args:
        json_file_path (str): JSON文件路径
        excel_file_path (str): 输出的Excel文件路径
    """
    
    # 定义所有可能的字段
    required_fields = [
        'prompt_id',
        'prompt', 
        'response_reference',
        'judge_prompt_template',
        'judge_system_prompt',
        'label'
    ]
    
    optional_fields = [
        'wind_ticker',
        'akshare_ticker', 
        'ground_truth',
        'time',
        'response_reference_translate',
        'yfinance_ticker'
    ]
    
    all_fields = required_fields + optional_fields
    
    try:
        # 读取JSON文件
        print(f"正在读取JSON文件: {json_file_path}")
        with open(json_file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        print(f"成功读取 {len(data)} 条记录")
        
        # 创建数据列表
        processed_data = []
        
        for i, item in enumerate(data):
            row = {}
            
            # 处理必需字段
            for field in required_fields:
                row[field] = item.get(field, '')
            
            # 处理可选字段
            for field in optional_fields:
                row[field] = item.get(field, '')
            
            processed_data.append(row)
            
            # 显示处理进度
            if (i + 1) % 100 == 0:
                print(f"已处理 {i + 1} 条记录...")
        
        # 创建DataFrame
        df = pd.DataFrame(processed_data, columns=all_fields)
        
        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "FinSearchComp数据"
        
        # 设置标题样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # 写入数据到工作表
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # 设置标题行样式
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # 设置合理的列宽范围
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 设置数据行样式
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        
        # 保存Excel文件
        wb.save(excel_file_path)
        print(f"Excel文件已成功保存到: {excel_file_path}")
        
        # 显示统计信息
        print(f"\n数据统计:")
        print(f"总记录数: {len(processed_data)}")
        print(f"总字段数: {len(all_fields)}")
        print(f"必需字段: {len(required_fields)}")
        print(f"可选字段: {len(optional_fields)}")
        
        # 显示可选字段的填充情况
        print(f"\n可选字段填充情况:")
        for field in optional_fields:
            filled_count = sum(1 for row in processed_data if row[field])
            fill_rate = (filled_count / len(processed_data)) * 100
            print(f"  {field}: {filled_count}/{len(processed_data)} ({fill_rate:.1f}%)")
        
        return True
        
    except FileNotFoundError:
        print(f"错误: 找不到JSON文件 {json_file_path}")
        return False
    except json.JSONDecodeError as e:
        print(f"错误: JSON文件格式错误 - {e}")
        return False
    except Exception as e:
        print(f"错误: {e}")
        return False

def main():
    """主函数"""
    # 设置文件路径
    json_file = "finsearchcomp_data.json"
    excel_file = "finsearchcomp_data.xlsx"
    
    # 检查JSON文件是否存在
    if not os.path.exists(json_file):
        print(f"错误: JSON文件 '{json_file}' 不存在")
        print("请确保JSON文件在当前目录下")
        return
    
    print("=" * 60)
    print("JSON转Excel工具")
    print("=" * 60)
    
    # 执行转换
    success = json_to_excel(json_file, excel_file)
    
    if success:
        print(f"\n✅ 转换完成！")
        print(f"输出文件: {excel_file}")
        
        # 显示文件大小
        file_size = os.path.getsize(excel_file)
        if file_size > 1024 * 1024:
            print(f"文件大小: {file_size / (1024 * 1024):.2f} MB")
        else:
            print(f"文件大小: {file_size / 1024:.2f} KB")
    else:
        print("\n❌ 转换失败！")

if __name__ == "__main__":
    main()